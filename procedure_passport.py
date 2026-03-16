import streamlit as st
import time
import pandas as pd
import uuid
import datetime
import io
import base64
import json
import gspread
from gspread_dataframe import get_as_dataframe, set_with_dataframe
from google.oauth2.service_account import Credentials
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

st.set_page_config(
    page_title="Procedure Passport",
    page_icon="🩺",
    layout="wide",
)

# ─────────────────────────────────────────────
# QUERY PARAMS  (magic link routing)
# ─────────────────────────────────────────────
query_params = st.query_params

# Only auto-route on the first load; once submitted we stay on the confirmation page.
if (
    query_params.get("mode") == "attending"
    and st.session_state.get("page", "login") not in ("attending_confirmation",)
    and not st.session_state.get("_magic_routed")
):
    st.session_state["page"]           = "attending_assessment"
    st.session_state["resident"]       = query_params.get("resident", "")
    st.session_state["procedure_id"]   = query_params.get("procedure_id", "")
    st.session_state["specialty_id"]   = query_params.get("specialty_id", "")
    st.session_state["attending_name"] = query_params.get("attending_name", "")
    st.session_state["_magic_routed"]  = True

# ─────────────────────────────────────────────
# SESSION STATE DEFAULTS
# ─────────────────────────────────────────────
_defaults: dict = {
    "page":                    "login",
    "resident":                None,
    "resident_name":           "",
    "scores":                  {},
    "date":                    datetime.date.today(),
    "notes":                   "",
    "current_case_id":         None,
    "attending_submission":    None,   # filled after magic-link submit
}
for _k, _v in _defaults.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
ADMINS = ["pjenkins9@gmail.com"]

RATING_OPTIONS = ["Not Assessed", "Not Done", "Not Yet", "Steer", "Prompt", "Back up", "Auto"]
RATING_TO_NUM  = {
    "Not Assessed": -1,
    "Not Done":      0,
    "Not Yet":       1,
    "Steer":         2,
    "Prompt":        3,
    "Back up":       4,
    "Auto":          5,
}
RATING_HEX = {
    "Not Assessed": "#E0E0E0",
    "Not Done":     "#D3D3D3",
    "Not Yet":      "#FF4D4D",
    "Steer":        "#FF944D",
    "Prompt":       "#FFD633",
    "Back up":      "#99E699",
    "Auto":         "#33CC33",
}
RATING_COLOR = {
    k: f"background-color:{v}; color:{'white' if k in ('Not Yet','Auto') else 'black'};"
    for k, v in RATING_HEX.items()
}

COMPLEXITY_HEX = {
    "Straight Forward": "#C8E6C9",
    "Moderate":         "#FFF59D",
    "Complex":          "#FFAB91",
}
O_SCORE_HEX = {
    "1": "#FF4D4D",
    "2": "#FF944D",
    "3": "#FFD633",
    "4": "#99E699",
    "5": "#33CC33",
}
O_SCORE_OPTIONS = [
    "— Make a selection —",
    "1 - Not Yet",
    "2 - Steer",
    "3 - Prompt",
    "4 - Backup",
    "5 - Auto",
]

SHEET_RESIDENTS  = "residents"
SHEET_ATTENDINGS = "attendings"
SHEET_PROCEDURES = "procedures"
SHEET_STEPS      = "steps"
SHEET_CASES      = "cases"
SHEET_SCORES     = "scores"
SHEET_SPECIALTY  = "specialties"

# ─────────────────────────────────────────────
# GOOGLE SHEETS HELPERS
# ─────────────────────────────────────────────

@st.cache_resource(show_spinner=False)
def get_gs_client():
    """Authorized gspread client — cached for the entire app session."""
    svc_json = json.loads(base64.b64decode(st.secrets["GOOGLE_SVC_B64"]).decode())
    creds = Credentials.from_service_account_info(
        svc_json,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    return gspread.authorize(creds)


def get_sheet(sheet_name: str):
    """Return a gspread worksheet, creating it if missing."""
    try:
        gc = get_gs_client()
        sh = gc.open_by_key(st.secrets["GOOGLE_SHEET_KEY"])
        try:
            return sh.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            return sh.add_worksheet(title=sheet_name, rows="500", cols="26")
    except Exception as exc:
        raise ConnectionError(f"Cannot reach Google Sheets: {exc}") from exc


@st.cache_data(ttl=60, show_spinner=False)
def read_sheet_df(sheet_name: str, expected_cols=None) -> pd.DataFrame:
    """Cached worksheet read (60 s TTL).  Returns empty DF if sheet is blank."""
    ws  = get_sheet(sheet_name)
    df  = get_as_dataframe(ws, evaluate_formulas=True, header=0)
    df  = df.dropna(how="all")
    if df.empty and expected_cols:
        return pd.DataFrame(columns=expected_cols)
    if expected_cols:
        for col in expected_cols:
            if col not in df.columns:
                df[col] = pd.NA
        df = df[expected_cols]
    return df


def write_sheet_df(sheet_name: str, df: pd.DataFrame) -> None:
    """Overwrite a worksheet then clear all cached reads so the UI stays fresh."""
    ws = get_sheet(sheet_name)
    ws.clear()
    set_with_dataframe(ws, df, include_index=False, include_column_header=True)
    st.cache_data.clear()  # invalidate all read caches after every write


@st.cache_data(ttl=60, show_spinner=False)
def load_refs():
    """Load all reference tables in one shot (cached 60 s)."""
    def _safe(name, cols):
        try:
            return read_sheet_df(name, expected_cols=cols)
        except Exception:
            return pd.DataFrame(columns=cols)

    spec_df  = _safe(SHEET_SPECIALTY,  ["specialty_id",  "specialty_name"])
    proc_df  = _safe(SHEET_PROCEDURES, ["procedure_id",  "procedure_name", "specialty_id"])
    steps_df = _safe(SHEET_STEPS,      ["step_id",       "procedure_id",   "step_order", "step_name"])
    atnd_df  = _safe(SHEET_ATTENDINGS, ["attending_id",  "attending_name", "specialty_id", "email"])
    return spec_df, proc_df, steps_df, atnd_df


# ─────────────────────────────────────────────
# DATA MUTATION HELPERS
# ─────────────────────────────────────────────

def ensure_resident(email: str, name: str = "", specialty_id=None) -> None:
    cols = ["email", "name", "specialty_id", "created_at"]
    df   = read_sheet_df(SHEET_RESIDENTS, expected_cols=cols)
    if email not in df["email"].values:
        df = pd.concat([df, pd.DataFrame([{
            "email":        email,
            "name":         name,
            "specialty_id": specialty_id,
            "created_at":   datetime.datetime.utcnow().isoformat(),
        }])], ignore_index=True)
        write_sheet_df(SHEET_RESIDENTS, df)   # also clears cache


def ensure_attending(name: str, specialty_id: str, email: str = "") -> None:
    cols = ["attending_id", "attending_name", "specialty_id", "email"]
    df   = read_sheet_df(SHEET_ATTENDINGS, expected_cols=cols)
    if name not in df["attending_name"].values:
        att_id = "A_" + specialty_id + "_" + name.replace(" ", "_").upper()
        df = pd.concat([df, pd.DataFrame([{
            "attending_id":   att_id,
            "attending_name": name,
            "specialty_id":   specialty_id,
            "email":          email,
        }])], ignore_index=True)
        write_sheet_df(SHEET_ATTENDINGS, df)


def ensure_procedure(proc_id: str, proc_name: str, specialty_id: str, steps_list: list) -> None:
    proc_cols = ["procedure_id", "procedure_name", "specialty_id"]
    procs_df  = read_sheet_df(SHEET_PROCEDURES, expected_cols=proc_cols)
    if proc_id not in procs_df["procedure_id"].values:
        procs_df = pd.concat([procs_df, pd.DataFrame([{
            "procedure_id":   proc_id,
            "procedure_name": proc_name,
            "specialty_id":   specialty_id,
        }])], ignore_index=True)
        write_sheet_df(SHEET_PROCEDURES, procs_df)

    step_cols = ["step_id", "procedure_id", "step_order", "step_name"]
    steps_df  = read_sheet_df(SHEET_STEPS, expected_cols=step_cols)
    if not (steps_df["procedure_id"] == proc_id).any():
        new_steps = pd.DataFrame([{
            "step_id":      f"S_{proc_id}_{i+1:02d}",
            "procedure_id": proc_id,
            "step_order":   i + 1,
            "step_name":    step,
        } for i, step in enumerate(steps_list)])
        steps_df = pd.concat([steps_df, new_steps], ignore_index=True)
        write_sheet_df(SHEET_STEPS, steps_df)


def save_case(
    resident_email: str,
    date,
    specialty_id: str,
    procedure_id: str,
    attending_id: str,
    scores_dict: dict,
    notes: str = "",
    case_complexity=None,
    overall_performance=None,
) -> str:
    """Persist a case + its step scores; returns the new case_id."""
    case_id   = uuid.uuid4().hex[:12]

    case_cols = ["case_id", "resident_email", "date", "specialty_id",
                 "procedure_id", "attending_id", "notes",
                 "case_complexity", "overall_performance"]
    cases_df  = read_sheet_df(SHEET_CASES, expected_cols=case_cols)
    cases_df  = pd.concat([cases_df, pd.DataFrame([{
        "case_id":             case_id,
        "resident_email":      resident_email,
        "date":                str(date),
        "specialty_id":        specialty_id,
        "procedure_id":        procedure_id,
        "attending_id":        attending_id,
        "notes":               notes,
        "case_complexity":     case_complexity,
        "overall_performance": overall_performance,
    }])], ignore_index=True)
    write_sheet_df(SHEET_CASES, cases_df)  # clears cache

    score_cols = ["case_id", "step_id", "rating", "rating_num",
                  "case_complexity", "overall_performance"]
    scores_df  = read_sheet_df(SHEET_SCORES, expected_cols=score_cols)
    new_rows   = [{
        "case_id":             case_id,
        "step_id":             step_id,
        "rating":              rating,
        "rating_num":          RATING_TO_NUM.get(rating),
        "case_complexity":     case_complexity,
        "overall_performance": overall_performance,
    } for step_id, rating in scores_dict.items()]
    scores_df  = pd.concat([scores_df, pd.DataFrame(new_rows)], ignore_index=True)
    write_sheet_df(SHEET_SCORES, scores_df)  # clears cache

    return case_id


# ─────────────────────────────────────────────
# STYLING HELPERS
# ─────────────────────────────────────────────

def style_df(df: pd.DataFrame, col: str):
    return df.style.map(lambda v: RATING_COLOR.get(v, ""), subset=[col])


def attending_display_name(attending_id: str, atnds_lookup: dict) -> str:
    """Resolve a display name from an attending_id, including magic_ IDs."""
    if attending_id in atnds_lookup:
        return atnds_lookup[attending_id]
    if isinstance(attending_id, str) and attending_id.startswith("magic_"):
        return attending_id[len("magic_"):].replace("_", " ")
    return attending_id or "Unknown"


def show_gs_error(exc: Exception) -> None:
    st.error(
        "⚠️ **Could not reach Google Sheets.** "
        "Check your network connection or try refreshing the page.\n\n"
        f"_Details: {exc}_"
    )


# ─────────────────────────────────────────────
# NAV HELPER
# ─────────────────────────────────────────────
def go_to(page: str) -> None:
    st.session_state["page"] = page
    st.rerun()


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
st.sidebar.title("🩺 Procedure Passport")

_logged_in = st.session_state.get("resident")
if _logged_in in ADMINS:
    if st.sidebar.button("⚙️ Admin Panel"):
        go_to("admin")

if _logged_in and st.session_state["page"] not in ("login", "attending_assessment", "attending_confirmation"):
    st.sidebar.markdown(f"👤 **{st.session_state.get('resident_name', '')}**")
    st.sidebar.markdown(f"_{_logged_in}_")
    st.sidebar.markdown("---")
    if st.sidebar.button("🚪 Logout"):
        for _k in list(st.session_state.keys()):
            del st.session_state[_k]
        st.cache_data.clear()
        st.rerun()

# ─────────────────────────────────────────────
# SHARED CSS
# ─────────────────────────────────────────────
st.markdown(
    """
<style>
/* Card-style sections */
.pp-card {
    background: var(--secondary-background-color);
    border-radius: 10px;
    padding: 1.2rem 1.5rem;
    margin-bottom: 1rem;
}
/* Pill badge */
.pp-badge {
    display: inline-block;
    border-radius: 12px;
    padding: 2px 10px;
    font-size: 0.82rem;
    font-weight: 600;
    margin: 2px;
}
/* Legend row */
.legend-row {
    display: flex;
    gap: 1rem;
    flex-wrap: wrap;
    align-items: center;
    margin-bottom: 0.5rem;
}
.legend-item {
    display: inline-flex;
    align-items: center;
    gap: 0.35rem;
    font-size: 0.85rem;
}
.legend-swatch {
    width: 14px;
    height: 14px;
    border-radius: 3px;
    border: 1px solid var(--secondary-background-color);
    display: inline-block;
}
</style>
""",
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────
# PAGE ROUTER
# ─────────────────────────────────────────────
page = st.session_state["page"]


# ════════════════════════════════════════════════════════════
# PAGE: LOGIN
# ════════════════════════════════════════════════════════════
if page == "login":
    col_c, col_r = st.columns([1, 1])
    with col_c:
        st.markdown("# 🩺 Procedure Passport")
        st.markdown("_Track your surgical skills journey, one procedure at a time._")
        st.markdown("---")
        email = st.text_input("Email address", placeholder="you@hospital.org")

        if st.button("Login →", use_container_width=True, type="primary"):
            if not email.strip():
                st.error("Please enter your email address.")
            else:
                try:
                    residents = read_sheet_df(
                        SHEET_RESIDENTS,
                        expected_cols=["email", "name", "specialty_id", "created_at"],
                    )
                    if email in ADMINS:
                        st.session_state.update(
                            resident=email, resident_name="Admin", page="admin"
                        )
                        st.rerun()
                    elif email in residents["email"].values:
                        row = residents.loc[residents["email"] == email].iloc[0]
                        st.session_state.update(
                            resident=email,
                            resident_name=row["name"],
                            specialty_id=row["specialty_id"],
                            page="home",
                        )
                        st.rerun()
                    else:
                        st.error("❌ Email not recognised. Ask an admin to add you.")
                except ConnectionError as exc:
                    show_gs_error(exc)


# ════════════════════════════════════════════════════════════
# PAGE: ADMIN PANEL
# ════════════════════════════════════════════════════════════
elif page == "admin":
    st.title("⚙️ Admin Panel")

    if st.button("🔄 Reload Data"):
        st.cache_data.clear()
        st.rerun()

    # ── Specialties ──────────────────────────────────────
    st.subheader("Specialties")
    try:
        specialties = read_sheet_df(SHEET_SPECIALTY, expected_cols=["specialty_id", "specialty_name"])
        st.dataframe(specialties, use_container_width=True)

        with st.expander("➕ Add Specialty"):
            new_spec_id   = st.text_input("Specialty ID (e.g., GS)")
            new_spec_name = st.text_input("Specialty name (e.g., General Surgery)")
            if st.button("Add Specialty", key="btn_add_spec"):
                if new_spec_id and new_spec_name:
                    if new_spec_id in specialties["specialty_id"].values:
                        st.warning("That ID already exists.")
                    else:
                        specialties = pd.concat(
                            [specialties, pd.DataFrame([{"specialty_id": new_spec_id,
                                                          "specialty_name": new_spec_name}])],
                            ignore_index=True,
                        )
                        write_sheet_df(SHEET_SPECIALTY, specialties)
                        st.success(f"✅ Added {new_spec_name}")
                        time.sleep(0.5)
                        st.rerun()
                else:
                    st.error("Please fill in both fields.")
    except ConnectionError as exc:
        show_gs_error(exc)

    st.markdown("---")

    # ── Residents ────────────────────────────────────────
    st.subheader("Residents")
    try:
        spec_df = read_sheet_df(SHEET_SPECIALTY, expected_cols=["specialty_id", "specialty_name"])
        spec_name_to_id = dict(zip(spec_df["specialty_name"], spec_df["specialty_id"]))

        residents = read_sheet_df(
            SHEET_RESIDENTS, expected_cols=["email", "name", "specialty_id", "created_at"]
        )
        disp = residents.merge(spec_df, how="left", on="specialty_id")
        st.dataframe(disp[["email", "name", "specialty_name", "created_at"]], use_container_width=True)

        with st.expander("➕ Add Resident"):
            new_res_email = st.text_input("Email")
            new_res_name  = st.text_input("Full name")
            new_res_spec  = st.selectbox("Specialty", list(spec_name_to_id.keys()), key="add_res_spec")
            if st.button("Add Resident", key="btn_add_res"):
                if new_res_email and new_res_name and new_res_spec:
                    ensure_resident(new_res_email, new_res_name, spec_name_to_id[new_res_spec])
                    st.success(f"✅ Added {new_res_email}")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.warning("Please fill in all fields.")

        if not residents.empty:
            with st.expander("🗑️ Delete Resident"):
                del_email = st.selectbox("Select resident to delete", residents["email"], key="del_res")
                if st.button("Delete", key="btn_del_res"):
                    updated = residents[residents["email"] != del_email].reset_index(drop=True)
                    write_sheet_df(SHEET_RESIDENTS, updated)
                    st.success(f"Deleted {del_email}")
                    time.sleep(0.5)
                    st.rerun()
    except ConnectionError as exc:
        show_gs_error(exc)

    st.markdown("---")

    # ── Attendings ───────────────────────────────────────
    st.subheader("Attendings")
    try:
        attendings = read_sheet_df(
            SHEET_ATTENDINGS, expected_cols=["attending_id", "attending_name", "specialty_id", "email"]
        )
        spec_df, _, _, _ = load_refs()
        st.dataframe(attendings, use_container_width=True)

        with st.expander("➕ Add Attending"):
            new_att_name  = st.text_input("Attending name")
            new_att_spec  = st.selectbox("Specialty", spec_df["specialty_name"], key="add_att_spec")
            new_att_email = st.text_input("Email (optional)")
            if st.button("Add Attending", key="btn_add_att"):
                if new_att_name:
                    spec_id = spec_df.loc[spec_df["specialty_name"] == new_att_spec, "specialty_id"].values[0]
                    ensure_attending(new_att_name, spec_id, new_att_email)
                    st.success(f"✅ Added {new_att_name}")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.error("Please enter an attending name.")

        if not attendings.empty:
            with st.expander("🗑️ Delete Attending"):
                del_att = st.selectbox("Select attending to delete", attendings["attending_name"], key="del_att")
                if st.button("Delete", key="btn_del_att"):
                    updated = attendings[attendings["attending_name"] != del_att].reset_index(drop=True)
                    write_sheet_df(SHEET_ATTENDINGS, updated)
                    st.success(f"Deleted {del_att}")
                    time.sleep(0.5)
                    st.rerun()
    except ConnectionError as exc:
        show_gs_error(exc)

    st.markdown("---")

    # ── Procedures ───────────────────────────────────────
    st.subheader("Procedures")
    try:
        spec_df, _, _, _ = load_refs()

        with st.expander("➕ Add New Procedure"):
            new_proc_id   = st.text_input("Procedure ID (e.g., CSEC)").strip().upper()
            new_proc_name = st.text_input("Procedure name (e.g., Cesarean Section)")
            new_proc_spec = st.selectbox("Specialty", spec_df["specialty_name"], key="add_proc_spec")
            steps_raw     = st.text_area("Steps (one per line)")
            new_steps     = [s.strip() for s in steps_raw.split("\n") if s.strip()]
            if st.button("Add Procedure", key="btn_add_proc"):
                if new_proc_id and new_proc_name and new_steps:
                    spec_id = spec_df.loc[spec_df["specialty_name"] == new_proc_spec, "specialty_id"].values[0]
                    ensure_procedure(new_proc_id, new_proc_name, spec_id, new_steps)
                    st.success(f"✅ Added {new_proc_name}")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.error("Please fill in all fields and at least one step.")

        with st.expander("✏️ Edit Existing Procedure"):
            procs_df = read_sheet_df(SHEET_PROCEDURES, expected_cols=["procedure_id", "procedure_name", "specialty_id"])
            if procs_df.empty:
                st.info("No procedures yet.")
            else:
                edit_proc    = st.selectbox("Select procedure", procs_df["procedure_name"], key="edit_proc_sel")
                sel_proc_id  = procs_df.loc[procs_df["procedure_name"] == edit_proc, "procedure_id"].values[0]
                new_pname    = st.text_input("Updated name", value=edit_proc, key="edit_proc_name")
                new_steps_ra = st.text_area("Updated steps (blank = keep current)", key="edit_proc_steps")
                new_edit_stp = [s.strip() for s in new_steps_ra.split("\n") if s.strip()]

                if st.button("Update Procedure", key="btn_upd_proc"):
                    procs_df.loc[procs_df["procedure_id"] == sel_proc_id, "procedure_name"] = new_pname
                    write_sheet_df(SHEET_PROCEDURES, procs_df)
                    if new_edit_stp:
                        steps_df = read_sheet_df(
                            SHEET_STEPS, expected_cols=["step_id", "procedure_id", "step_order", "step_name"]
                        )
                        steps_df = steps_df[steps_df["procedure_id"] != sel_proc_id]
                        updated_steps = pd.DataFrame([{
                            "step_id":      f"S_{sel_proc_id}_{i+1:02d}",
                            "procedure_id": sel_proc_id,
                            "step_order":   i + 1,
                            "step_name":    s,
                        } for i, s in enumerate(new_edit_stp)])
                        steps_df = pd.concat([steps_df, updated_steps], ignore_index=True)
                        write_sheet_df(SHEET_STEPS, steps_df)
                    st.success(f"✅ Updated '{new_pname}'")
                    time.sleep(0.5)
                    st.rerun()
    except ConnectionError as exc:
        show_gs_error(exc)

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("⬅️ Back to Login"):
            go_to("login")
    with col2:
        if st.button("🏠 Resident Home"):
            go_to("home")


# ════════════════════════════════════════════════════════════
# PAGE: HOME
# ════════════════════════════════════════════════════════════
elif page == "home":
    st.title(f"👋 Welcome back, {st.session_state['resident_name']}")
    st.markdown("_What would you like to do today?_")
    st.markdown("")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown('<div class="pp-card">', unsafe_allow_html=True)
        st.markdown("### ➕ New Assessment")
        st.markdown("Start a new procedure case and record step ratings.")
        if st.button("Start Assessment", use_container_width=True, type="primary"):
            go_to("start")
        st.markdown("</div>", unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="pp-card">', unsafe_allow_html=True)
        st.markdown("### 📊 Cumulative Dashboard")
        st.markdown("View your progress heatmap and learning curve over time.")
        if st.button("View Dashboard", use_container_width=True):
            go_to("cumulative")
        st.markdown("</div>", unsafe_allow_html=True)

    with c3:
        st.markdown('<div class="pp-card">', unsafe_allow_html=True)
        st.markdown("### 💬 Comments")
        st.markdown("Browse and export all attending feedback.")
        if st.button("View Comments", use_container_width=True):
            go_to("comments")
        st.markdown("</div>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════
# PAGE: START CASE
# ════════════════════════════════════════════════════════════
elif page == "start":
    st.title("📋 Start Assessment")

    try:
        spec_df, proc_df, steps_df, atnd_df = load_refs()
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    spec_map   = dict(zip(spec_df["specialty_name"], spec_df["specialty_id"]))
    id_to_name = dict(zip(spec_df["specialty_id"],   spec_df["specialty_name"]))
    is_admin   = st.session_state["resident"] in ADMINS

    if is_admin:
        selected_spec_name = st.selectbox("Specialty", list(spec_map.keys()))
        specialty_id       = spec_map[selected_spec_name]
        st.session_state["specialty_id"] = specialty_id
    else:
        specialty_id       = st.session_state.get("specialty_id")
        selected_spec_name = id_to_name.get(specialty_id, "Unknown Specialty")
        st.markdown(f"**Specialty:** {selected_spec_name}")
        if specialty_id is None:
            st.error("No specialty assigned. Contact an admin.")
            st.stop()

    procs = proc_df[proc_df["specialty_id"] == specialty_id]
    atnds = atnd_df[atnd_df["specialty_id"] == specialty_id]

    if procs.empty:
        st.warning("⚠️ No procedures configured for this specialty.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()
    if atnds.empty:
        st.warning("⚠️ No attendings configured for this specialty.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    proc_map = dict(zip(procs["procedure_name"], procs["procedure_id"]))
    atnd_map = dict(zip(atnds["attending_name"], atnds["attending_id"]))

    procedure = st.selectbox("Procedure", list(proc_map.keys()))
    attending = st.selectbox("Attending",  list(atnd_map.keys()))
    case_date = st.date_input("Date", st.session_state["date"])

    st.session_state["procedure_id"] = proc_map[procedure]
    st.session_state["attending_id"] = atnd_map[attending]
    st.session_state["date"]         = case_date

    # ── Magic link for attending ──────────────────────────
    if not is_admin:
        safe_att  = atnds.loc[atnds["attending_id"] == st.session_state["attending_id"],
                               "attending_name"].values[0].replace(" ", "_")
        base_url  = "https://procedurepassport.streamlit.app"
        magic_url = (
            f"{base_url}/?mode=attending"
            f"&resident={st.session_state['resident']}"
            f"&procedure_id={st.session_state['procedure_id']}"
            f"&specialty_id={specialty_id}"
            f"&attending_name={safe_att}"
        )
        with st.expander("🔗 Magic Link for Attending (click to expand)", expanded=False):
            st.markdown(
                "Share this link with your attending so they can submit their evaluation directly:"
            )
            st.code(magic_url, language="text")

    st.markdown("---")
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        if st.button("⬅️ Back to Home"):
            go_to("home")
    with col3:
        if st.button("Start Assessment →", type="primary", use_container_width=True):
            st.session_state["scores"] = {}
            st.session_state["notes"]  = ""
            go_to("assessment")


# ════════════════════════════════════════════════════════════
# PAGE: ASSESSMENT
# ════════════════════════════════════════════════════════════
elif page == "assessment":
    try:
        _, _, steps_df, _ = load_refs()
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Start"):
            go_to("start")
        st.stop()

    steps = steps_df[steps_df["procedure_id"] == st.session_state["procedure_id"]].sort_values("step_order")
    if steps.empty:
        st.error("No steps defined for this procedure. Ask an admin to add steps.")
        if st.button("⬅️ Back to Start"):
            go_to("start")
        st.stop()

    st.title("📝 Assessment")

    st.session_state["case_complexity"] = st.selectbox(
        "Case Complexity",
        ["Straight Forward", "Moderate", "Complex"],
        index=["Straight Forward", "Moderate", "Complex"].index(
            st.session_state.get("case_complexity", "Straight Forward")
        ),
    )

    st.markdown("#### Step-Level Ratings")

    if st.button("↺ Mark All as 'Not Assessed'"):
        for _, row in steps.iterrows():
            st.session_state["scores"][row["step_id"]] = "Not Assessed"

    for _, row in steps.iterrows():
        step_id   = row["step_id"]
        step_name = row["step_name"]
        current   = st.session_state["scores"].get(step_id, "Not Assessed")
        st.session_state["scores"][step_id] = st.selectbox(
            step_name,
            RATING_OPTIONS,
            index=RATING_OPTIONS.index(current) if current in RATING_OPTIONS else 0,
            key=f"score_{step_id}",
        )

    current_o = st.session_state.get("overall_performance", O_SCORE_OPTIONS[0])
    st.session_state["overall_performance"] = st.selectbox(
        "Overall Performance (O-Score)",
        O_SCORE_OPTIONS,
        index=O_SCORE_OPTIONS.index(current_o) if current_o in O_SCORE_OPTIONS else 0,
    )

    st.session_state["notes"] = st.text_area("Comments / Feedback", st.session_state.get("notes", ""))

    if all(v == "Not Assessed" for v in st.session_state["scores"].values()):
        st.warning("⚠️ All steps are marked 'Not Assessed'.")

    st.markdown("---")
    col1, col2 = st.columns([1, 2])
    with col1:
        if st.button("⬅️ Back to Start"):
            go_to("start")
    with col2:
        if st.button("Finish & Save →", type="primary", use_container_width=True):
            if st.session_state["overall_performance"] == O_SCORE_OPTIONS[0]:
                st.warning("Please select an Overall Performance rating.")
            else:
                try:
                    st.session_state["current_case_id"] = save_case(
                        resident_email=st.session_state["resident"],
                        date=st.session_state["date"],
                        specialty_id=st.session_state["specialty_id"],
                        procedure_id=st.session_state["procedure_id"],
                        attending_id=st.session_state["attending_id"],
                        scores_dict=st.session_state["scores"],
                        case_complexity=st.session_state["case_complexity"],
                        overall_performance=st.session_state["overall_performance"],
                        notes=st.session_state.get("notes", ""),
                    )
                    go_to("dashboard")
                except ConnectionError as exc:
                    show_gs_error(exc)


# ════════════════════════════════════════════════════════════
# PAGE: SINGLE-CASE DASHBOARD
# ════════════════════════════════════════════════════════════
elif page == "dashboard":
    try:
        _, _, steps_df, _ = load_refs()
    except ConnectionError as exc:
        show_gs_error(exc)
        st.stop()

    steps = steps_df[steps_df["procedure_id"] == st.session_state["procedure_id"]].sort_values("step_order")

    st.title("✅ Case Saved")
    st.success(f"Case ID: `{st.session_state.get('current_case_id', '—')}`")

    data = [{"Step": row["step_name"],
             "Rating": st.session_state["scores"].get(row["step_id"], "")}
            for _, row in steps.iterrows()]
    df   = pd.DataFrame(data)
    st.dataframe(style_df(df, "Rating"), use_container_width=True)

    meta_col1, meta_col2 = st.columns(2)
    with meta_col1:
        st.markdown(f"**Case Complexity:** {st.session_state.get('case_complexity', '—')}")
    with meta_col2:
        st.markdown(f"**Overall Performance:** {st.session_state.get('overall_performance', '—')}")

    if st.session_state.get("notes", "").strip():
        st.markdown("**Comments:**")
        st.info(st.session_state["notes"])

    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("⬅️ Back to Assessment"):
            go_to("assessment")
    with col2:
        if st.button("🏠 Home"):
            go_to("home")
    with col3:
        if st.button("➕ New Assessment", type="primary"):
            go_to("start")


# ════════════════════════════════════════════════════════════
# PAGE: COMMENTS DASHBOARD
# ════════════════════════════════════════════════════════════
elif page == "comments":
    st.title("💬 Comments Dashboard")
    resident = st.session_state.get("resident")
    if not resident:
        st.error("Not logged in.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    try:
        cases_df = read_sheet_df(
            SHEET_CASES,
            expected_cols=["case_id", "resident_email", "date", "specialty_id",
                           "procedure_id", "attending_id", "notes",
                           "case_complexity", "overall_performance"],
        )
        procs_df = read_sheet_df(SHEET_PROCEDURES, expected_cols=["procedure_id", "procedure_name", "specialty_id"])
        atnds_df = read_sheet_df(SHEET_ATTENDINGS, expected_cols=["attending_id", "attending_name", "specialty_id", "email"])
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    # Deduplicate cases to prevent a fan-out if the sheet has duplicate rows.
    cases_df = cases_df.drop_duplicates(subset=["case_id"])

    res_cases = cases_df[cases_df["resident_email"] == resident].copy()
    res_cases["notes"] = res_cases["notes"].fillna("").astype(str)
    res_cases = res_cases[res_cases["notes"].str.strip() != ""]

    if res_cases.empty:
        st.info("No comments recorded yet.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
    else:
        # Resolve attending names — magic_ IDs never appear in the attendings sheet,
        # so we decode them directly from the ID string instead of joining.
        atnds_lookup = dict(zip(atnds_df["attending_id"], atnds_df["attending_name"]))
        res_cases["attending_name"] = res_cases["attending_id"].apply(
            lambda aid: attending_display_name(str(aid), atnds_lookup)
        )

        # Deduplicate procs so a fanout can't multiply rows.
        procs_dedup = procs_df.drop_duplicates(subset=["procedure_id"])
        merged = res_cases.merge(procs_dedup[["procedure_id", "procedure_name"]], on="procedure_id", how="left")
        merged = merged.rename(columns={
            "date":                "Date",
            "procedure_name":      "Procedure",
            "attending_name":      "Attending",
            "case_complexity":     "Case Complexity",
            "overall_performance": "Overall Performance",
            "notes":               "Comments",
        })
        merged = merged[["Date", "Procedure", "Attending",
                          "Case Complexity", "Overall Performance", "Comments"]].sort_values("Date", ascending=False)

        st.dataframe(merged, use_container_width=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            merged.to_excel(writer, index=False, sheet_name="Comments")
        st.download_button(
            label="📥 Download as Excel",
            data=output.getvalue(),
            file_name=f"{resident}_comments.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if st.button("⬅️ Back to Home"):
            go_to("home")


# ════════════════════════════════════════════════════════════
# PAGE: CUMULATIVE DASHBOARD
# ════════════════════════════════════════════════════════════
elif page == "cumulative":
    st.title("📊 Cumulative Dashboard")
    resident = st.session_state.get("resident")
    if not resident:
        st.error("Not logged in.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    try:
        cases_df  = read_sheet_df(SHEET_CASES,  expected_cols=["case_id", "resident_email", "date",
                                                                "specialty_id", "procedure_id",
                                                                "attending_id", "notes",
                                                                "case_complexity", "overall_performance"])
        scores_df = read_sheet_df(SHEET_SCORES, expected_cols=["case_id", "step_id", "rating", "rating_num",
                                                                "case_complexity", "overall_performance"])
        steps_df  = read_sheet_df(SHEET_STEPS,  expected_cols=["step_id", "procedure_id", "step_order", "step_name"])
        procs_df  = read_sheet_df(SHEET_PROCEDURES, expected_cols=["procedure_id", "procedure_name", "specialty_id"])
        atnds_df  = read_sheet_df(SHEET_ATTENDINGS, expected_cols=["attending_id", "attending_name", "specialty_id", "email"])
    except ConnectionError as exc:
        show_gs_error(exc)
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    for col in ["case_complexity", "overall_performance"]:
        if col not in scores_df.columns:
            scores_df[col] = pd.NA

    # Deduplicate all lookup tables before any join so duplicate sheet rows
    # can't produce phantom extra rows in the dashboard.
    cases_df  = cases_df.drop_duplicates(subset=["case_id"])
    scores_df = scores_df.drop_duplicates(subset=["case_id", "step_id"])
    steps_df  = steps_df.drop_duplicates(subset=["step_id"])

    res_cases = cases_df[cases_df["resident_email"] == resident]
    if res_cases.empty:
        st.info("No cases logged yet.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    # Resolve attending names (handle magic_ IDs)
    atnds_lookup = dict(zip(atnds_df["attending_id"], atnds_df["attending_name"]))
    res_cases    = res_cases.copy()
    res_cases["attending_name"] = res_cases["attending_id"].apply(
        lambda aid: attending_display_name(str(aid), atnds_lookup)
    )

    res_cases_small = res_cases[["case_id", "date", "procedure_id",
                                  "attending_name", "case_complexity", "overall_performance"]].copy()
    res_cases_small = res_cases_small.rename(columns={"procedure_id": "case_procedure_id"})

    steps_small = steps_df[["step_id", "procedure_id", "step_name", "step_order"]].rename(
        columns={"procedure_id": "step_procedure_id"}
    )

    procs_map = procs_df.set_index("procedure_id")["procedure_name"].to_dict()

    merged = (
        scores_df[["case_id", "step_id", "rating", "rating_num"]]
        .merge(res_cases_small, on="case_id", how="inner")
        .merge(steps_small, on="step_id", how="left")
    )

    if merged.empty:
        st.info("No assessment data yet.")
        if st.button("⬅️ Back to Home"):
            go_to("home")
        st.stop()

    # ── Procedure selector ────────────────────────────────
    proc_ids      = merged["case_procedure_id"].dropna().unique()
    selected_proc = st.selectbox(
        "Procedure",
        options=list(proc_ids),
        format_func=lambda x: procs_map.get(x, x),
    )

    proc_data = merged[merged["case_procedure_id"] == selected_proc].copy()
    ordered_steps = (
        steps_df[steps_df["procedure_id"] == selected_proc]
        .sort_values("step_order")["step_name"]
        .tolist()
    )

    # ── Pivot for heatmap ─────────────────────────────────
    pivot = proc_data.pivot_table(
        index=["date", "attending_name", "case_id", "case_complexity", "overall_performance"],
        columns="step_name",
        values="rating",
        aggfunc="first",
    ).reset_index()

    for step in ordered_steps:
        if step not in pivot.columns:
            pivot[step] = pd.NA

    pivot = pivot[["date", "attending_name", "case_id", "case_complexity", "overall_performance"] + ordered_steps]

    # ── Screenshot-friendly heatmap ───────────────────────
    st.markdown(
        "### Progress Heatmap\n"
        "Most recent cases at the top. Zoom out to screenshot this grid. 📸"
    )

    screenshot_df = pivot.sort_values("date", ascending=False).drop(columns=["case_id"])
    all_cols      = list(screenshot_df.columns)

    def _color_step(val):
        if pd.isna(val):
            return ""
        return f"background-color: {RATING_HEX.get(val, '')}"

    def _color_complexity(val):
        if pd.isna(val):
            return ""
        return f"background-color: {COMPLEXITY_HEX.get(val, '')}"

    def _color_o_score(val):
        if not isinstance(val, str):
            return ""
        key = val.split("-")[0].strip()
        return f"background-color: {O_SCORE_HEX.get(key, '')}"

    styled = screenshot_df.style
    if ordered_steps:
        styled = styled.map(_color_step, subset=ordered_steps)
    styled = (
        styled
        .map(_color_complexity, subset=["case_complexity"])
        .map(_color_o_score,    subset=["overall_performance"])
        .hide(axis="index")
        .set_properties(
            subset=["date", "attending_name"],
            **{"min-width": "120px", "white-space": "nowrap"},
        )
        .set_properties(
            subset=["case_complexity", "overall_performance"],
            **{"min-width": "90px", "text-align": "center"},
        )
    )
    if ordered_steps:
        styled = styled.set_properties(
            subset=ordered_steps,
            **{"min-width": "40px", "max-width": "40px", "text-align": "center"},
        )

    table_styles = [
        {"selector": "table",       "props": [("border-collapse", "collapse"), ("margin", "0 auto")]},
        {"selector": "th, td",      "props": [("border", "1px solid var(--secondary-background-color)"),
                                               ("padding", "4px"), ("font-size", "0.8rem")]},
        {"selector": "th.col_heading", "props": [("text-align", "center"), ("vertical-align", "bottom"),
                                                   ("font-weight", "600")]},
    ]
    for idx, col_name in enumerate(all_cols):
        if col_name in ordered_steps:
            table_styles.append({
                "selector": f"th.col_heading.level0.col{idx}",
                "props": [("writing-mode", "vertical-rl"), ("text-orientation", "mixed"),
                           ("white-space", "nowrap"), ("font-size", "0.75rem"), ("padding", "4px 2px")],
            })

    comp_idx = all_cols.index("case_complexity")
    o_idx    = all_cols.index("overall_performance")
    table_styles += [
        {"selector": f"td.col{comp_idx}", "props": [("color", "transparent")]},
        {"selector": f"td.col{o_idx}",    "props": [("color", "transparent")]},
    ]

    styled = styled.set_table_styles(table_styles)
    st.markdown(styled.to_html(), unsafe_allow_html=True)

    # ── Legends ───────────────────────────────────────────
    def _swatch(color, label):
        return (
            f'<span class="legend-item">'
            f'<span class="legend-swatch" style="background-color:{color};"></span>{label}'
            f'</span>'
        )

    st.markdown("#### Ratings Legend")
    st.markdown(
        '<div class="legend-row">' +
        "".join(_swatch(v, k) for k, v in RATING_HEX.items()) +
        "</div>",
        unsafe_allow_html=True,
    )

    st.markdown("#### Case Complexity")
    st.markdown(
        '<div class="legend-row">' +
        "".join(_swatch(v, k) for k, v in COMPLEXITY_HEX.items()) +
        "</div>",
        unsafe_allow_html=True,
    )

    st.markdown("#### O-Score")
    o_labels = {"1": "1–Not Yet", "2": "2–Steer", "3": "3–Prompt", "4": "4–Backup", "5": "5–Auto"}
    st.markdown(
        '<div class="legend-row">' +
        "".join(_swatch(v, o_labels[k]) for k, v in O_SCORE_HEX.items()) +
        "</div>",
        unsafe_allow_html=True,
    )

    # ── Learning Curve Chart ──────────────────────────────
    st.markdown("---")
    st.markdown("### 📈 Learning Curve — Rating Over Time per Step")

    # Build a time-ordered table: case# (x), step, avg rating_num
    curve_data = (
        proc_data[["case_id", "date", "step_name", "rating_num"]]
        .dropna(subset=["rating_num"])
        .copy()
    )
    curve_data["rating_num"] = pd.to_numeric(curve_data["rating_num"], errors="coerce")
    curve_data = curve_data.dropna(subset=["rating_num"])
    curve_data = curve_data[curve_data["rating_num"] >= 0]   # exclude Not Assessed (-1)

    if curve_data.empty:
        st.info("Not enough rated steps yet to draw a learning curve.")
    else:
        # Assign a sequential case number per date
        case_order = (
            proc_data[["case_id", "date"]]
            .drop_duplicates()
            .sort_values("date")
            .reset_index(drop=True)
        )
        case_order["case_num"] = range(1, len(case_order) + 1)
        curve_data = curve_data.merge(case_order[["case_id", "case_num"]], on="case_id", how="left")

        step_avg = (
            curve_data.groupby(["case_num", "step_name"])["rating_num"]
            .mean()
            .reset_index()
        )

        # Show only up to the 8 most commonly assessed steps to keep the chart readable
        top_steps = (
            curve_data.groupby("step_name")["rating_num"].count()
            .sort_values(ascending=False)
            .head(8)
            .index.tolist()
        )
        step_avg  = step_avg[step_avg["step_name"].isin(top_steps)]

        fig, ax = plt.subplots(figsize=(10, 4))
        fig.patch.set_facecolor("none")
        ax.set_facecolor("none")

        cmap   = plt.get_cmap("tab10")
        colors = [cmap(i) for i in range(len(top_steps))]

        for i, step_name in enumerate(top_steps):
            sdata = step_avg[step_avg["step_name"] == step_name].sort_values("case_num")
            ax.plot(sdata["case_num"], sdata["rating_num"],
                    marker="o", linewidth=1.8, markersize=4,
                    color=colors[i], label=step_name)

        ax.set_xlabel("Case Number", fontsize=9)
        ax.set_ylabel("Avg Rating (0=Not Done → 5=Auto)", fontsize=9)
        ax.set_yticks(range(6))
        ax.set_yticklabels(["Not Done", "Not Yet", "Steer", "Prompt", "Back up", "Auto"], fontsize=7)
        ax.set_ylim(-0.3, 5.3)
        ax.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax.legend(loc="lower right", fontsize=7, ncol=2)
        ax.grid(axis="y", alpha=0.25)
        ax.spines[["top", "right"]].set_visible(False)

        st.pyplot(fig, transparent=True)
        plt.close(fig)

    # ── Excel export ──────────────────────────────────────
    st.markdown("---")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pivot.to_excel(writer, index=False, sheet_name="Cumulative")
        ws_xl = writer.sheets["Cumulative"]
        from openpyxl.styles import PatternFill, Font

        step_fill_map = {k: v.lstrip("#") for k, v in RATING_HEX.items() if k not in ("Not Assessed",)}
        step_fill_map["Not Assessed"] = "E0E0E0"

        start_col = 6
        for xl_row in ws_xl.iter_rows(
            min_row=2, max_row=ws_xl.max_row,
            min_col=start_col, max_col=5 + len(ordered_steps),
        ):
            for cell in xl_row:
                val = cell.value
                if val in step_fill_map:
                    cell.fill = PatternFill(
                        start_color=step_fill_map[val],
                        end_color=step_fill_map[val],
                        fill_type="solid",
                    )
                    cell.font = Font(color="FFFFFF" if val in ("Not Yet", "Auto") else "000000")

    st.download_button(
        label=f"📥 Download Excel — {procs_map.get(selected_proc, selected_proc)}",
        data=output.getvalue(),
        file_name=f"{resident}_{selected_proc}_cumulative.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if st.button("⬅️ Back to Home"):
        go_to("home")


# ════════════════════════════════════════════════════════════
# PAGE: ATTENDING ASSESSMENT (magic link)
# ════════════════════════════════════════════════════════════
elif page == "attending_assessment":
    resident_email = st.session_state.get("resident", "")
    procedure_id   = st.session_state.get("procedure_id", "")
    specialty_id   = st.session_state.get("specialty_id", "")
    attending_name = st.session_state.get("attending_name", "Unknown")

    if not (resident_email and procedure_id and specialty_id):
        st.error("⚠️ Missing required information in this link. Please ask the resident to resend.")
        st.stop()

    # Decode URL-safe attending name
    display_attending = attending_name.replace("_", " ")

    st.title("📝 Attending Evaluation")
    st.markdown(
        f'<div class="pp-card">'
        f'<b>Resident:</b> {resident_email}<br>'
        f'<b>Procedure:</b> <code>{procedure_id}</code><br>'
        f'<b>Attending:</b> {display_attending}'
        f'</div>',
        unsafe_allow_html=True,
    )

    try:
        _, _, steps_df, _ = load_refs()
    except ConnectionError as exc:
        show_gs_error(exc)
        st.stop()

    steps = steps_df[steps_df["procedure_id"] == procedure_id].sort_values("step_order")
    if steps.empty:
        st.error("This procedure has no defined steps. Please contact the program coordinator.")
        st.stop()

    case_date       = st.date_input("Date of Procedure", value=datetime.date.today())
    case_complexity = st.selectbox("Case Complexity", ["Straight Forward", "Moderate", "Complex"])

    st.markdown("#### Step-Level Ratings")
    scores: dict = {}
    for _, row in steps.iterrows():
        step_id   = row["step_id"]
        step_name = row["step_name"]
        scores[step_id] = st.selectbox(
            step_name, RATING_OPTIONS, key=f"att_score_{step_id}"
        )

    o_score = st.selectbox("Overall Performance (O-Score)", O_SCORE_OPTIONS)
    notes   = st.text_area("Comments / Feedback (optional)")

    st.markdown("---")
    if st.button("✅ Submit Evaluation", type="primary", use_container_width=True):
        if o_score == O_SCORE_OPTIONS[0]:
            st.warning("Please select an Overall Performance rating before submitting.")
        else:
            try:
                case_id = save_case(
                    resident_email=resident_email,
                    date=case_date,
                    specialty_id=specialty_id,
                    procedure_id=procedure_id,
                    attending_id=f"magic_{attending_name}",   # magic_ prefix; decoded on display
                    scores_dict=scores,
                    notes=notes,
                    case_complexity=case_complexity,
                    overall_performance=o_score,
                )
                # Store submission summary for the confirmation page
                st.session_state["attending_submission"] = {
                    "case_id":             case_id,
                    "resident_email":      resident_email,
                    "procedure_id":        procedure_id,
                    "attending_name":      display_attending,
                    "date":                str(case_date),
                    "case_complexity":     case_complexity,
                    "overall_performance": o_score,
                    "notes":               notes,
                    "scores":              scores,
                    "steps":               steps[["step_id", "step_name"]].to_dict("records"),
                }
                go_to("attending_confirmation")
            except ConnectionError as exc:
                show_gs_error(exc)


# ════════════════════════════════════════════════════════════
# PAGE: ATTENDING CONFIRMATION
# ════════════════════════════════════════════════════════════
elif page == "attending_confirmation":
    sub = st.session_state.get("attending_submission")
    if not sub:
        st.error("No submission data found.")
        st.stop()

    st.title("✅ Evaluation Submitted")
    st.success("Thank you! Your evaluation has been recorded.")

    st.markdown(
        f'<div class="pp-card">'
        f'<b>Resident:</b> {sub["resident_email"]}<br>'
        f'<b>Attending:</b> {sub["attending_name"]}<br>'
        f'<b>Procedure:</b> <code>{sub["procedure_id"]}</code><br>'
        f'<b>Date:</b> {sub["date"]}<br>'
        f'<b>Case Complexity:</b> {sub["case_complexity"]}<br>'
        f'<b>Overall Performance:</b> {sub["overall_performance"]}<br>'
        f'<b>Case ID:</b> <code>{sub["case_id"]}</code>'
        f'</div>',
        unsafe_allow_html=True,
    )

    if sub["notes"].strip():
        st.markdown("**Comments submitted:**")
        st.info(sub["notes"])

    st.markdown("#### Step Ratings Submitted")
    step_rows = []
    for step_rec in sub["steps"]:
        step_id   = step_rec["step_id"]
        step_name = step_rec["step_name"]
        rating    = sub["scores"].get(step_id, "—")
        step_rows.append({"Step": step_name, "Rating": rating})

    summary_df = pd.DataFrame(step_rows)
    st.dataframe(style_df(summary_df, "Rating"), use_container_width=True)

    st.markdown("---")
    st.markdown("_You may now close this window. The resident can view the evaluation in their dashboard._")
